"""Generate cost-model.xlsx — the Local Dev Environment ROI spreadsheet.

This script produces an Excel workbook with live formulas (not hardcoded numbers),
so that anyone changing an Inputs cell sees the ROI Summary update automatically.

Usage:
    pip install openpyxl
    python cost-model.py

Output:
    cost-model.xlsx  (the workbook)
    cost-model.csv   (a flat CSV fallback of the inputs and key outputs)

Design notes
------------
- Every editable parameter lives on the `Inputs` tab and has a defined name
  (a "named range" in Excel). All other tabs reference parameters by name,
  so formulas are readable: e.g. `=team_size * hourly_rate * ...` rather
  than `=Inputs!B5 * Inputs!B6 * ...`.

- The model is CALIBRATED with real team data (Jan-May 2026, ~4.25 months):
  26 anonymised rework work-items across 8 developers, plus 24 measured
  build durations. See the `Real-Bug-Data` and `Real-Build-Times` tabs.

- The `use_bottom_up_multiplier` toggle on the Inputs tab switches between:
    TRUE  - development-stage cost is computed bottom-up from the
            Development-Fix-Workflow tab (recommended).
    FALSE - development-stage cost uses IBM's textbook 15x multiplier and
            adds the workflow tax as a separate line (avoids double-counting).

- The `use_real_data` toggle wires the model to the Real-Bug-Data and
  Real-Build-Times tabs (rework rate, avg PRs per rework item, median
  pipeline duration). Set to 0 to fall back to industry placeholders.

Environment topology assumed: Local -> Development -> Staging -> Production.
CI exists but does not gate deployment; CI-flagged bugs flow through to the
Development environment where they are operationally caught and fixed.

Rework workflow (from the team's documented 46-step process):
- Steps 1-35  = dev-test cycle (repeated when a bug is caught BEFORE QA handoff)
- Steps 36-40 = QA cycle       (repeated when a bug is caught BY QA, in addition to 1-35)
- Steps 41-46 = master merge   (executed ONCE per work-item; not repeated for rework)

So each rework cycle redoes ~35/46 (pre-QA) or ~40/46 (post-QA) of the process.
The blended redo factor is controlled by `pct_rework_caught_pre_qa` on Inputs.
"""
from __future__ import annotations

import csv
from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName


# ---------------------------------------------------------------------------
# Real team data (anonymised) — source: team-bug-data-raw.xlsx
# Period: Jan 1 2026 - May 8 2026 (~4.25 months).  Real names masked Dev1..Dev8.
# Each tuple is (developer_alias, work_item_id, pr_count).
# ---------------------------------------------------------------------------

REAL_BUG_DATA: list[tuple[str, int, int]] = [
    ("Dev1", 42851, 7),
    ("Dev1", 43844, 4),
    ("Dev1", 44413, 2),
    ("Dev1", 44858, 2),
    ("Dev1", 44975, 2),
    ("Dev1", 45977, 2),
    ("Dev1", 44858, 2),
    ("Dev2", 45225, 2),
    ("Dev2", 43666, 6),
    ("Dev2", 40653, 4),
    ("Dev2", 45225, 2),
    ("Dev2", 45751, 2),
    ("Dev3", 45750, 2),
    ("Dev3", 44848, 2),
    ("Dev4", 43694, 6),
    ("Dev4", 44032, 3),
    ("Dev4", 44146, 3),
    ("Dev4", 43403, 3),
    ("Dev4", 40471, 3),
    ("Dev4", 46247, 2),
    ("Dev5", 42666, 6),
    ("Dev5", 42665, 3),
    ("Dev5", 45195, 3),
    ("Dev5", 44045, 2),
    ("Dev7", 43668, 2),
    ("Dev8", 42665, 2),
]

# Real build durations from the team's Azure DevOps pipelines.
# (build_name, duration_minutes)
REAL_BUILD_TIMES: list[tuple[str, float]] = [
    ("Build Nuget Packages", 6),
    ("DripsBackend-PR-Build", 24),
    ("Lead Processor Build", 9),
    ("DripsBackend-Build", 10),
    ("Api.Gateway", 4),
    ("Api.Private", 4),
    ("Api.Private.Numbers", 5),
    ("Api.Public", 6),
    ("DripsBackend Repo Dev", 37),
    ("DripsBackend-Build-Dev", 22),
    ("Processors.OutboundSMS - Dev", 4),
    ("Services.Campaign - Dev", 2),
    ("Providers.Services - Dev", 2),
    ("Engine.Services - Dev", 2),
    ("DAL.Services - Dev", 3),
    ("Providers - Dev", 5),
    ("Hosting - Dev", 3),
    ("Engine - Dev", 3),
    ("Services - Dev", 4),
    ("Providers.Api.Bandwidth - Dev", 4),
    ("Number Inventory API - dev", 6),
    ("Common.Services - Dev", 2),
    ("Dal - Dev", 2),
    ("Processors.WebHooks", 5),
]

# The team's documented end-to-end "PR from local to development to master" process.
# Step index, description, repeat_status:
#   "once"      = executes once per work-item
#   "pre_qa"    = repeated each rework cycle when dev catches the bug pre-QA
#   "post_qa"   = repeated each rework cycle when QA catches the bug (in addition to pre_qa)
#   "master"    = master-merge tail, not repeated per rework cycle
PROCESS_STEPS: list[tuple[int, str, str]] = [
    (1,  "Pull latest from development", "pre_qa"),
    (2,  "Build local NuGet packages (PowerShell script)", "pre_qa"),
    (3,  "Create feature branch + make code changes", "pre_qa"),
    (4,  "Test locally using workarounds (or skip if not possible)", "pre_qa"),
    (5,  "Run GitHub Copilot PR code analysis", "pre_qa"),
    (6,  "Address issues from Copilot analysis", "pre_qa"),
    (7,  "Re-run Copilot PR analysis to confirm clean", "pre_qa"),
    (8,  "Push changes to remote", "pre_qa"),
    (9,  "Create new PR with Copilot analysis attached", "pre_qa"),
    (10, "Raise PR for approval + reviewer review", "pre_qa"),
    (11, "Adjust code based on reviewer feedback", "pre_qa"),
    (12, "Push additional changes + re-run Copilot analysis", "pre_qa"),
    (13, "Reviewer re-reviews", "pre_qa"),
    (14, "Approve + deploy to development environment", "pre_qa"),
    (15, "Wait for several CI/CD builds per PR", "pre_qa"),
    (16, "QA tests run automatically", "pre_qa"),
    (17, "Developer investigates any failed QA tests", "pre_qa"),
    (18, "Re-run failing QA test locally to confirm real vs false failure", "pre_qa"),
    (19, "If real bug -> begin rework cycle", "pre_qa"),
    (20, "Pull latest from development branch", "pre_qa"),
    (21, "Rebuild NuGet packages", "pre_qa"),
    (22, "Create new feature branch", "pre_qa"),
    (23, "Address identified bug", "pre_qa"),
    (24, "Test locally if possible", "pre_qa"),
    (25, "Run Copilot code analysis report", "pre_qa"),
    (26, "Address issues from Copilot report", "pre_qa"),
    (27, "Re-run Copilot analysis", "pre_qa"),
    (28, "Push changes back to remote", "pre_qa"),
    (29, "Create new PR copying Copilot analysis", "pre_qa"),
    (30, "Get approvers from other developers", "pre_qa"),
    (31, "Address any further developer feedback", "pre_qa"),
    (32, "Re-run Copilot analysis after fix", "pre_qa"),
    (33, "Get developers to re-approve", "pre_qa"),
    (34, "Approve PR + redeploy to development", "pre_qa"),
    (35, "Wait for builds to complete for deployment", "pre_qa"),
    (36, "Investigate any failing builds (QA test or real failure)", "post_qa"),
    (37, "QA resource starts manual feature testing", "post_qa"),
    (38, "If QA finds a bug not covered by tests -> new rework cycle", "post_qa"),
    (39, "Confirm card is clean / ready for master merge", "post_qa"),
    (40, "Plan master-merge steps", "post_qa"),
    (41, "Cherry-pick PR of changes into master branch (happy path)", "master"),
    (42, "If merge conflicts: pull latest from master + development", "master"),
    (43, "Create topic branch off master, cherry-pick from development", "master"),
    (44, "Address merge conflicts + push", "master"),
    (45, "Get approvers to approve PR into master", "master"),
    (46, "Merge to master (final)", "master"),
]


# ---------------------------------------------------------------------------
# Defaults — edit these to reflect your team, then re-run this script.
# ---------------------------------------------------------------------------

@dataclass
class Inputs:
    # Team and finance
    team_size: int = 8
    hourly_rate: float = 85.0  # fully-loaded $/hour
    working_days_per_year: int = 250
    working_days_per_sprint: int = 9  # 2-week sprints, 1 day for ceremonies
    sprints_per_year: int = 26

    # Bug volume and distribution (today)
    # Stages: Local -> Development -> Staging -> Production
    # (CI exists but does not gate deployment, so bugs flagged by CI still flow to Dev
    # and are operationally caught/fixed at the Dev stage.)
    bugs_per_year: int = 400
    pct_today_local: float = 0.35
    pct_today_dev: float = 0.45  # consolidates the typical CI-15% + Integration-30% since CI doesn't gate
    pct_today_stg: float = 0.15
    pct_today_prod: float = 0.05

    # Bug distribution (target after investment)
    pct_target_local: float = 0.75
    pct_target_dev: float = 0.18
    pct_target_stg: float = 0.05
    pct_target_prod: float = 0.02

    # Cost multipliers (IBM SSI baseline, adapted to 4-stage topology)
    mult_local: float = 1.0
    mult_dev: float = 15.0  # IBM "Integration" multiplier — same workflow, different name
    mult_stg: float = 40.0
    mult_prod: float = 100.0

    # Local-caught bug baseline cost
    local_fix_hours: float = 0.92  # ~55 minutes
    # local_fix_cost_usd derived: local_fix_hours * hourly_rate

    # Integration-fix workflow (13-step)
    triage_min: float = 10
    branch_repro_min: float = 20
    code_fix_min: float = 45
    ai_check_min: float = 7
    push_wait_min: float = 25
    push_idle_factor: float = 0.5  # half the wait is recovered productivity
    pr_create_min: float = 10
    review_wait_hours: float = 4.0  # wall-clock wait before reviewer picks up
    review_wait_idle_factor: float = 0.5
    reviewer_review_min: float = 30
    review_rounds: float = 1.5  # avg rounds per bug-fix PR
    review_round_dev_min: float = 30  # per round, dev side
    review_round_reviewer_min: float = 20  # per round, reviewer side
    approve_merge_min: float = 5
    redeploy_min: float = 20
    redeploy_idle_factor: float = 0.5
    reverify_min: float = 30
    retry_rate: float = 0.30
    ai_check_tool_cost_per_run: float = 0.30  # AI tool token cost in USD

    # Cycle time and sprint
    iterations_per_dev_per_day: int = 12
    cycle_time_today_min: float = 25.0
    cycle_time_target_min: float = 2.0
    cycle_idle_recovery_factor: float = 0.5
    context_switch_multiplier: float = 1.2

    # Pipeline cost
    builds_per_dev_per_day: int = 6
    pipeline_duration_min: float = 25.0
    queue_wait_min: float = 5.0
    pipeline_cost_per_min: float = 0.015  # marginal MS-hosted derivation
    parallel_jobs_extra: int = 5  # legacy single-bucket (used when use_real_data=0)
    parallel_job_monthly_cost: float = 40.0  # legacy avg cost
    ms_hosted_parallel_jobs: int = 5   # actual count of MS-hosted parallel jobs
    self_hosted_parallel_jobs: int = 50  # actual count of self-hosted agents
    ms_hosted_monthly_cost: float = 40.0   # $40/mo MS-hosted private
    self_hosted_monthly_cost: float = 15.0  # $15/mo self-hosted
    pipeline_failure_rate: float = 0.30
    pct_retries_caused_by_local: float = 0.50  # of failures, % preventable
    pipeline_idle_recovery_factor: float = 0.5

    # Workarounds (placeholder team total)
    workaround_hours_per_week_total: float = 12.0  # 8 devs x 1.5 hrs each
    workaround_weeks_per_year: int = 50

    # PCE & shift-left
    pct_env_or_integration_related: float = 0.35
    pct_preventable_with_local: float = 0.70

    # Investment side
    setup_engineer_hours: float = 60
    onboarding_doc_hours: float = 20
    training_hours_per_dev: float = 4
    machine_upgrade_cost: float = 6000.0
    docker_license_cost_per_user_per_month: float = 21.0
    apply_docker_license_cost: bool = True  # set False if <250-employee org
    annual_maintenance_hours: float = 100  # ~5% of one engineer

    # Real-data integration
    data_period_months: float = 4.25  # Jan 1 - May 8, 2026 in months
    total_work_items_in_period: int = 200  # placeholder — pull from ADO Boards
    pct_rework_caught_pre_qa: float = 0.60  # share of rework caught BEFORE QA handoff
    redo_factor_pre_qa: float = 35.0 / 46.0   # ~76% of process repeated per dev-caught rework
    redo_factor_post_qa: float = 40.0 / 46.0  # ~87% of process repeated per QA-caught rework

    # Strategy toggles
    use_bottom_up_multiplier: bool = True
    use_real_data: bool = True
    discount_rate: float = 0.10  # for 3-year NPV


# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
SECTION_FONT = Font(bold=True, size=11)
TOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")
TOTAL_FONT = Font(bold=True, size=11)
INPUT_FILL = PatternFill("solid", fgColor="FBE4D5")  # editable cells
NOTE_FONT = Font(italic=True, color="595959", size=9)


def header_row(ws, row: int, headers: list[str]):
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = BORDER


def section_row(ws, row: int, title: str, cols: int = 5):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = SECTION_FILL
    ws.cell(row=row, column=1, value=title).font = SECTION_FONT


def total_row(ws, row: int, cols: int = 5):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT


def fmt(ws, ranges: list[tuple[str, str]]):
    """Apply number format to ranges. ranges = [(cell_range, fmt_string), ...]."""
    for rng, format_str in ranges:
        for row in ws[rng]:
            for cell in row:
                cell.number_format = format_str


def widen(ws, widths: dict[str, int]):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def add_named(wb: Workbook, name: str, sheet: str, cell: str):
    """Create a workbook-scoped named range pointing at one cell."""
    dn = DefinedName(name=name, attr_text=f"{sheet}!${cell.split('$')[-1] if '$' in cell else cell.replace(chr(36), '').upper()}")
    # Simpler approach: build the absolute reference cleanly.
    col_letters = "".join(c for c in cell if c.isalpha())
    row_digits = "".join(c for c in cell if c.isdigit())
    ref = f"{sheet}!${col_letters}${row_digits}"
    dn = DefinedName(name=name, attr_text=ref)
    wb.defined_names[name] = dn


# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------

def build_workbook(inp: Inputs) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    build_inputs(wb, inp)
    build_real_bug_data(wb, inp)
    build_real_build_times(wb, inp)
    build_process_steps(wb)
    build_defect_distribution(wb)
    build_bug_cost_by_stage(wb)
    build_development_fix_workflow(wb)
    build_pipeline_cost(wb)
    build_cycle_time_sprint(wb)
    build_workarounds(wb)
    build_roi_summary(wb)
    build_sensitivity(wb)

    return wb


# ---------------------------------------------------------------------------
# Inputs sheet
# ---------------------------------------------------------------------------

def build_inputs(wb: Workbook, inp: Inputs):
    ws = wb.create_sheet("Inputs")
    widen(ws, {"A": 42, "B": 14, "C": 12, "D": 60})
    ws["A1"] = "Inputs — edit the orange cells; everything else recalculates"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")

    header_row(ws, 3, ["Parameter", "Value", "Unit", "Notes"])

    # Each tuple: (display_name, named_range, value, unit, notes)
    rows: list[tuple[str, str, object, str, str]] = [
        ("--- Team & Finance ---", None, "", "", ""),
        ("Team size (active developers)", "team_size", inp.team_size, "devs", "Active devs on this team"),
        ("Loaded hourly rate", "hourly_rate", inp.hourly_rate, "$/hr", "Fully loaded: salary + benefits + overhead"),
        ("Working days per year", "working_days_per_year", inp.working_days_per_year, "days", ""),
        ("Working days per sprint per dev", "working_days_per_sprint", inp.working_days_per_sprint, "days", "9 typical for 2-week sprints"),
        ("Sprints per year", "sprints_per_year", inp.sprints_per_year, "sprints", ""),

        ("--- Bug Volume & Distribution Today ---", None, "", "", ""),
        ("Bugs per year", "bugs_per_year", inp.bugs_per_year, "bugs", "Annualised from last 180 days of Azure Boards"),
        ("% bugs caught Local today", "pct_today_local", inp.pct_today_local, "%", "From WIQL query, Found In = Local"),
        ("% bugs caught Development today", "pct_today_dev", inp.pct_today_dev, "%", "CI doesn't gate so CI-flagged bugs roll into Dev"),
        ("% bugs caught Staging today", "pct_today_stg", inp.pct_today_stg, "%", ""),
        ("% bugs caught Production today", "pct_today_prod", inp.pct_today_prod, "%", ""),

        ("--- Target Distribution (Post-Investment) ---", None, "", "", ""),
        ("% bugs caught Local target", "pct_target_local", inp.pct_target_local, "%", "Capers Jones median-elite range"),
        ("% bugs caught Development target", "pct_target_dev", inp.pct_target_dev, "%", ""),
        ("% bugs caught Staging target", "pct_target_stg", inp.pct_target_stg, "%", ""),
        ("% bugs caught Production target", "pct_target_prod", inp.pct_target_prod, "%", ""),

        ("--- IBM Cost-of-Defect Multipliers ---", None, "", "", ""),
        ("Multiplier (Local)", "mult_local", inp.mult_local, "x", "Baseline"),
        ("Multiplier (Development)", "mult_dev", inp.mult_dev, "x", "IBM Integration 15x; toggle below replaces with bottom-up value"),
        ("Multiplier (Staging)", "mult_stg", inp.mult_stg, "x", "IBM SSI"),
        ("Multiplier (Production)", "mult_prod", inp.mult_prod, "x", "IBM SSI / NIST RTI 2002"),
        ("Local fix hours (baseline)", "local_fix_hours", inp.local_fix_hours, "hrs", "Time to fix a bug found locally"),

        ("--- Development-Fix Workflow ---", None, "", "", ""),
        ("Triage / work item creation", "triage_min", inp.triage_min, "min", ""),
        ("Branch creation + repro attempt", "branch_repro_min", inp.branch_repro_min, "min", ""),
        ("Code fix", "code_fix_min", inp.code_fix_min, "min", ""),
        ("AI code-check prompt run", "ai_check_min", inp.ai_check_min, "min", ""),
        ("AI tool cost per run", "ai_check_tool_cost", inp.ai_check_tool_cost_per_run, "$", "Token cost"),
        ("Push wait (build + tests)", "push_wait_min", inp.push_wait_min, "min", ""),
        ("Push wait idle recovery factor", "push_idle_factor", inp.push_idle_factor, "frac", "0=fully idle, 1=fully recovered"),
        ("PR create + write description", "pr_create_min", inp.pr_create_min, "min", ""),
        ("Reviewer wait (wall-clock hours)", "review_wait_hours", inp.review_wait_hours, "hrs", "DORA review wait time"),
        ("Reviewer wait idle recovery factor", "review_wait_idle_factor", inp.review_wait_idle_factor, "frac", ""),
        ("Reviewer review time per pass", "reviewer_review_min", inp.reviewer_review_min, "min", "Reviewer side"),
        ("Average review rounds", "review_rounds", inp.review_rounds, "rounds", "Total iterations per PR"),
        ("Review round dev-side time", "review_round_dev_min", inp.review_round_dev_min, "min/round", ""),
        ("Review round reviewer-side time", "review_round_reviewer_min", inp.review_round_reviewer_min, "min/round", ""),
        ("Approve + merge", "approve_merge_min", inp.approve_merge_min, "min", ""),
        ("Redeploy to integration", "redeploy_min", inp.redeploy_min, "min", "Mostly idle wait"),
        ("Redeploy idle recovery factor", "redeploy_idle_factor", inp.redeploy_idle_factor, "frac", ""),
        ("Re-run regression + re-verify", "reverify_min", inp.reverify_min, "min", ""),
        ("Retry rate (1 + this multiplies total)", "retry_rate", inp.retry_rate, "frac", "30% typical bug-fix PR retry rate"),

        ("--- Cycle Time & Sprint ---", None, "", "", ""),
        ("Inner-loop iterations per dev per day", "iterations_per_dev_per_day", inp.iterations_per_dev_per_day, "iters", "Microsoft inner-loop research: 8-15"),
        ("Cycle time today (min/iteration)", "cycle_time_today_min", inp.cycle_time_today_min, "min", "Median; replace from time-tracking week"),
        ("Cycle time target (min/iteration)", "cycle_time_target_min", inp.cycle_time_target_min, "min", "Healthy local loop"),
        ("Cycle time idle recovery factor", "cycle_idle_recovery_factor", inp.cycle_idle_recovery_factor, "frac", ""),
        ("Context-switch multiplier", "context_switch_multiplier", inp.context_switch_multiplier, "x", "DeMarco/Weinberg 1.2-1.4"),

        ("--- Pipeline (ADO) ---", None, "", "", ""),
        ("Builds per dev per day", "builds_per_dev_per_day", inp.builds_per_dev_per_day, "builds", ""),
        ("Pipeline duration (min/run) — placeholder", "pipeline_duration_min", inp.pipeline_duration_min, "min", "Overridden by real_median_build_min when use_real_data=1"),
        ("Queue wait (min/run)", "queue_wait_min", inp.queue_wait_min, "min", "Pre-start"),
        ("Pipeline compute cost per min", "pipeline_cost_per_min", inp.pipeline_cost_per_min, "$/min", "MS-hosted derivation"),
        ("Extra parallel jobs (legacy single bucket)", "parallel_jobs_extra", inp.parallel_jobs_extra, "jobs", "Used only if use_real_data=0"),
        ("Parallel job monthly cost (legacy avg)", "parallel_job_monthly_cost", inp.parallel_job_monthly_cost, "$/mo", "Used only if use_real_data=0"),
        ("MS-hosted parallel jobs", "ms_hosted_parallel_jobs", inp.ms_hosted_parallel_jobs, "jobs", "Real count from team — 5"),
        ("Self-hosted parallel jobs", "self_hosted_parallel_jobs", inp.self_hosted_parallel_jobs, "jobs", "Real count from team — ~50"),
        ("MS-hosted monthly cost", "ms_hosted_monthly_cost", inp.ms_hosted_monthly_cost, "$/mo", "$40/mo per MS-hosted private job"),
        ("Self-hosted monthly cost", "self_hosted_monthly_cost", inp.self_hosted_monthly_cost, "$/mo", "~$15/mo amortised hardware + power"),
        ("Pipeline failure rate", "pipeline_failure_rate", inp.pipeline_failure_rate, "frac", "DORA change-failure rate"),
        ("% of failures caused by missing local env", "pct_retries_caused_by_local", inp.pct_retries_caused_by_local, "frac", "Recoverable share"),
        ("Pipeline idle recovery factor", "pipeline_idle_recovery_factor", inp.pipeline_idle_recovery_factor, "frac", ""),

        ("--- Workarounds ---", None, "", "", ""),
        ("Total team workaround hours per week", "workaround_hours_per_week_total", inp.workaround_hours_per_week_total, "hrs/wk", "From self-assessment in developer-workarounds.md"),
        ("Working weeks per year", "workaround_weeks_per_year", inp.workaround_weeks_per_year, "weeks", ""),

        ("--- Shift-Left Allocation ---", None, "", "", ""),
        ("% bugs env/integration-related", "pct_env_or_integration_related", inp.pct_env_or_integration_related, "frac", "DORA range 25-45%"),
        ("% preventable with proper local env", "pct_preventable_with_local", inp.pct_preventable_with_local, "frac", "Conservative 70% default"),

        ("--- Investment ---", None, "", "", ""),
        ("Setup engineer hours (one-time)", "setup_engineer_hours", inp.setup_engineer_hours, "hrs", "Build docker-compose, validate emulators"),
        ("Onboarding/doc hours (one-time)", "onboarding_doc_hours", inp.onboarding_doc_hours, "hrs", ""),
        ("Training hours per dev", "training_hours_per_dev", inp.training_hours_per_dev, "hrs/dev", ""),
        ("Machine upgrade cost (one-time, total)", "machine_upgrade_cost", inp.machine_upgrade_cost, "$", "RAM/SSD if needed"),
        ("Docker license $ per user per month", "docker_license_cost_per_user_per_month", inp.docker_license_cost_per_user_per_month, "$/user/mo", "$21 Business; $0 if <250-employee org"),
        ("Apply Docker license cost? (1=yes, 0=no)", "apply_docker_license_cost", 1 if inp.apply_docker_license_cost else 0, "0/1", "0 for small org or free alternatives"),
        ("Annual maintenance hours", "annual_maintenance_hours", inp.annual_maintenance_hours, "hrs/yr", "~5% of one engineer"),

        ("--- Real-Data Calibration (Jan-May 2026) ---", None, "", "", ""),
        ("Data period (months)", "data_period_months", inp.data_period_months, "months", "Jan 1 - May 8 2026 = ~4.25 mo"),
        ("Total work items in period (est.)", "total_work_items_in_period", inp.total_work_items_in_period, "items", "Pull from ADO Boards for this period"),
        ("% of rework caught pre-QA-handoff", "pct_rework_caught_pre_qa", inp.pct_rework_caught_pre_qa, "frac", "Remainder caught by QA; both redo most of the 46-step process"),
        ("Redo factor — dev-caught (steps 1-35 of 46)", "redo_factor_pre_qa", inp.redo_factor_pre_qa, "frac", "Pre-QA bugs repeat ~76% of the process"),
        ("Redo factor — QA-caught (steps 1-40 of 46)", "redo_factor_post_qa", inp.redo_factor_post_qa, "frac", "Post-QA bugs repeat ~87% of the process"),

        ("--- Strategy Toggles ---", None, "", "", ""),
        ("Use bottom-up multiplier (1=yes, 0=no)", "use_bottom_up_multiplier", 1 if inp.use_bottom_up_multiplier else 0, "0/1", "1 = compute integration multiplier from workflow tab"),
        ("Use real team data (1=yes, 0=no)", "use_real_data", 1 if inp.use_real_data else 0, "0/1", "1 = source pipeline duration, rework rate, PR cost from real tabs"),
        ("Discount rate for NPV", "discount_rate", inp.discount_rate, "frac", "10% default"),
    ]

    row = 4
    for display, name, value, unit, notes in rows:
        if name is None:
            section_row(ws, row, display, cols=4)
        else:
            ws.cell(row=row, column=1, value=display).border = BORDER
            cell = ws.cell(row=row, column=2, value=value)
            cell.fill = INPUT_FILL
            cell.font = Font(bold=True)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="right")
            if isinstance(value, float):
                if "frac" in unit or "%" in unit:
                    cell.number_format = "0.0%"
                elif "$" in unit:
                    cell.number_format = "$#,##0.00"
                else:
                    cell.number_format = "#,##0.00"
            elif isinstance(value, int):
                cell.number_format = "#,##0"
            ws.cell(row=row, column=3, value=unit).border = BORDER
            note_cell = ws.cell(row=row, column=4, value=notes)
            note_cell.font = NOTE_FONT
            note_cell.border = BORDER
            note_cell.alignment = Alignment(wrap_text=True)

            # Define name pointing at this cell
            ref = f"Inputs!$B${row}"
            dn = DefinedName(name=name, attr_text=ref)
            wb.defined_names[name] = dn
        row += 1


# ---------------------------------------------------------------------------
# Real-Bug-Data sheet — anonymised rework items + summary stats
# ---------------------------------------------------------------------------

def build_real_bug_data(wb: Workbook, inp: Inputs):
    ws = wb.create_sheet("Real-Bug-Data")
    widen(ws, {"A": 12, "B": 14, "C": 14, "D": 56})

    ws["A1"] = "Real Rework Data (anonymised) — Jan 1 to May 8, 2026"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")
    ws["A2"] = "Source: team-bug-data-raw.xlsx (kept outside the repo). Developer names masked Dev1..Dev8."
    ws["A2"].font = NOTE_FONT
    ws.merge_cells("A2:D2")

    header_row(ws, 4, ["Developer", "Work Item", "PR Count", "Notes"])

    row = 5
    first_data_row = row
    for dev, wi, pr_count in REAL_BUG_DATA:
        ws.cell(row=row, column=1, value=dev).border = BORDER
        ws.cell(row=row, column=2, value=wi).border = BORDER
        c = ws.cell(row=row, column=3, value=pr_count)
        c.border = BORDER
        c.number_format = "#,##0"
        ws.cell(row=row, column=4, value="").border = BORDER
        row += 1
    last_data_row = row - 1

    row += 1
    section_row(ws, row, "Summary Statistics (rework items only)", cols=4)
    row += 1
    ws.cell(row=row, column=1, value="Rework items observed").border = BORDER
    ws.cell(row=row, column=3, value=f"=COUNTA(B{first_data_row}:B{last_data_row})").number_format = "#,##0"
    items_cell = f"C{row}"
    wb.defined_names["real_rework_items_count"] = DefinedName(
        name="real_rework_items_count",
        attr_text=f"'Real-Bug-Data'!$C${row}",
    )
    row += 1
    ws.cell(row=row, column=1, value="Total PRs across rework items").border = BORDER
    ws.cell(row=row, column=3, value=f"=SUM(C{first_data_row}:C{last_data_row})").number_format = "#,##0"
    total_prs_cell = f"C{row}"
    wb.defined_names["real_rework_prs_total"] = DefinedName(
        name="real_rework_prs_total",
        attr_text=f"'Real-Bug-Data'!$C${row}",
    )
    row += 1
    ws.cell(row=row, column=1, value="Mean PRs per rework item").border = BORDER
    ws.cell(row=row, column=3, value=f"={total_prs_cell}/{items_cell}").number_format = "0.00"
    avg_prs_cell = f"C{row}"
    wb.defined_names["real_avg_prs_per_rework"] = DefinedName(
        name="real_avg_prs_per_rework",
        attr_text=f"'Real-Bug-Data'!$C${row}",
    )
    row += 1
    ws.cell(row=row, column=1, value="Median PRs per rework item").border = BORDER
    ws.cell(row=row, column=3, value=f"=MEDIAN(C{first_data_row}:C{last_data_row})").number_format = "0.00"
    row += 1
    ws.cell(row=row, column=1, value="Max PRs per rework item").border = BORDER
    ws.cell(row=row, column=3, value=f"=MAX(C{first_data_row}:C{last_data_row})").number_format = "#,##0"
    row += 1
    ws.cell(row=row, column=1, value="Min PRs per rework item").border = BORDER
    ws.cell(row=row, column=3, value=f"=MIN(C{first_data_row}:C{last_data_row})").number_format = "#,##0"
    row += 2

    section_row(ws, row, "Annualised Projection", cols=4)
    row += 1
    ws.cell(row=row, column=1, value="Annualised rework items per year").border = BORDER
    ws.cell(row=row, column=3, value=f"={items_cell}*12/data_period_months").number_format = "#,##0"
    annual_rework_cell = f"C{row}"
    wb.defined_names["real_rework_items_annual"] = DefinedName(
        name="real_rework_items_annual",
        attr_text=f"'Real-Bug-Data'!$C${row}",
    )
    row += 1
    ws.cell(row=row, column=1, value="Annualised extra PR cycles (rework only)").border = BORDER
    ws.cell(row=row, column=3, value=f"={annual_rework_cell}*({avg_prs_cell}-1)").number_format = "#,##0"
    extra_prs_cell = f"C{row}"
    wb.defined_names["real_extra_pr_cycles_annual"] = DefinedName(
        name="real_extra_pr_cycles_annual",
        attr_text=f"'Real-Bug-Data'!$C${row}",
    )
    row += 1
    ws.cell(row=row, column=1, value="Rework rate (% of work items needing rework)").border = BORDER
    ws.cell(row=row, column=3, value=f"={items_cell}/total_work_items_in_period").number_format = "0.0%"
    row += 2

    section_row(ws, row, "Blended Redo Factor", cols=4)
    row += 1
    ws.cell(row=row, column=1, value="Blended redo factor (pre-QA share × pre-QA + post-QA share × post-QA)").border = BORDER
    ws.cell(
        row=row,
        column=3,
        value="=pct_rework_caught_pre_qa*redo_factor_pre_qa+(1-pct_rework_caught_pre_qa)*redo_factor_post_qa",
    ).number_format = "0.000"
    wb.defined_names["redo_factor_blended"] = DefinedName(
        name="redo_factor_blended",
        attr_text=f"'Real-Bug-Data'!$C${row}",
    )
    row += 2

    section_row(ws, row, "Per-Developer Breakdown", cols=4)
    row += 1
    header_row(ws, row, ["Developer", "Rework items", "Sum PRs", "Avg PRs/item"])
    row += 1
    devs = sorted({d for d, _, _ in REAL_BUG_DATA})
    for dev in devs:
        ws.cell(row=row, column=1, value=dev).border = BORDER
        ws.cell(row=row, column=2, value=f'=COUNTIF(A{first_data_row}:A{last_data_row},"{dev}")').number_format = "#,##0"
        ws.cell(row=row, column=3, value=f'=SUMIF(A{first_data_row}:A{last_data_row},"{dev}",C{first_data_row}:C{last_data_row})').number_format = "#,##0"
        ws.cell(row=row, column=4, value=f"=IF(B{row}=0,0,C{row}/B{row})").number_format = "0.00"
        for col in (3, 4):
            ws.cell(row=row, column=col).border = BORDER
        ws.cell(row=row, column=2).border = BORDER
        row += 1


# ---------------------------------------------------------------------------
# Real-Build-Times sheet — measured ADO pipeline durations
# ---------------------------------------------------------------------------

def build_real_build_times(wb: Workbook, inp: Inputs):
    ws = wb.create_sheet("Real-Build-Times")
    widen(ws, {"A": 38, "B": 16, "C": 50})

    ws["A1"] = "Real Pipeline Build Durations (measured)"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:C1")
    ws["A2"] = "Source: team-bug-data-raw.xlsx, 'Builds' tab. Most components run 2-6 min; PR-build and dev-deploy dominate."
    ws["A2"].font = NOTE_FONT
    ws.merge_cells("A2:C2")

    header_row(ws, 4, ["Build name", "Duration (min)", "Notes"])

    row = 5
    first_row = row
    for name, mins in REAL_BUILD_TIMES:
        ws.cell(row=row, column=1, value=name).border = BORDER
        c = ws.cell(row=row, column=2, value=mins)
        c.number_format = "0.0"
        c.border = BORDER
        ws.cell(row=row, column=3, value="").border = BORDER
        row += 1
    last_row = row - 1

    row += 1
    section_row(ws, row, "Summary Statistics", cols=3)
    row += 1
    ws.cell(row=row, column=1, value="Build count").border = BORDER
    ws.cell(row=row, column=2, value=f"=COUNT(B{first_row}:B{last_row})").number_format = "#,##0"
    row += 1
    ws.cell(row=row, column=1, value="Median build duration (min)").border = BORDER
    ws.cell(row=row, column=2, value=f"=MEDIAN(B{first_row}:B{last_row})").number_format = "0.0"
    median_cell = f"B{row}"
    wb.defined_names["real_median_build_min"] = DefinedName(
        name="real_median_build_min",
        attr_text=f"'Real-Build-Times'!$B${row}",
    )
    row += 1
    ws.cell(row=row, column=1, value="Mean build duration (min)").border = BORDER
    ws.cell(row=row, column=2, value=f"=AVERAGE(B{first_row}:B{last_row})").number_format = "0.0"
    wb.defined_names["real_mean_build_min"] = DefinedName(
        name="real_mean_build_min",
        attr_text=f"'Real-Build-Times'!$B${row}",
    )
    row += 1
    ws.cell(row=row, column=1, value="P95 build duration (min)").border = BORDER
    ws.cell(row=row, column=2, value=f"=PERCENTILE.INC(B{first_row}:B{last_row},0.95)").number_format = "0.0"
    row += 1
    ws.cell(row=row, column=1, value="Max build duration (min)").border = BORDER
    ws.cell(row=row, column=2, value=f"=MAX(B{first_row}:B{last_row})").number_format = "0.0"
    wb.defined_names["real_max_build_min"] = DefinedName(
        name="real_max_build_min",
        attr_text=f"'Real-Build-Times'!$B${row}",
    )
    row += 1
    ws.cell(row=row, column=1, value="Min build duration (min)").border = BORDER
    ws.cell(row=row, column=2, value=f"=MIN(B{first_row}:B{last_row})").number_format = "0.0"
    row += 1
    ws.cell(row=row, column=1, value="Sum of all builds (worst-case per PR)").border = BORDER
    ws.cell(row=row, column=2, value=f"=SUM(B{first_row}:B{last_row})").number_format = "0.0"
    row += 1
    ws.cell(
        row=row,
        column=1,
        value="Critical-path build (max of PR-build + dev-deploy; bottleneck per PR)",
    ).border = BORDER
    # The two longest builds are typically the PR-build and the deploy-to-dev
    # build. Approximate critical path as the largest individual build (since
    # most others run in parallel within their stage).
    ws.cell(row=row, column=2, value=f"=MAX(B{first_row}:B{last_row})").number_format = "0.0"
    wb.defined_names["real_critical_path_min"] = DefinedName(
        name="real_critical_path_min",
        attr_text=f"'Real-Build-Times'!$B${row}",
    )
    row += 2

    section_row(ws, row, "Effective Pipeline Duration (used by model)", cols=3)
    row += 1
    ws.cell(row=row, column=1, value="Effective pipeline duration").border = BORDER
    ws.cell(
        row=row,
        column=2,
        value="=IF(use_real_data=1,real_critical_path_min,pipeline_duration_min)",
    ).number_format = "0.0"
    ws.cell(
        row=row,
        column=3,
        value=(
            "Critical-path build duration (the longest single build the developer "
            "must wait for) when use_real_data=1; placeholder otherwise. "
            "Median/mean are far too optimistic for wall-clock wait per PR."
        ),
    ).font = NOTE_FONT
    wb.defined_names["effective_pipeline_min"] = DefinedName(
        name="effective_pipeline_min",
        attr_text=f"'Real-Build-Times'!$B${row}",
    )


# ---------------------------------------------------------------------------
# Process-Steps sheet — the team's documented 46-step process
# ---------------------------------------------------------------------------

def build_process_steps(wb: Workbook):
    ws = wb.create_sheet("Process-Steps")
    widen(ws, {"A": 6, "B": 70, "C": 18, "D": 22})

    ws["A1"] = "Team's Documented 46-Step Process — and What Repeats on Rework"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")
    ws["A2"] = (
        "Steps 1-35 are repeated when a developer catches a bug pre-QA-handoff. "
        "Steps 36-40 are additionally repeated when QA finds the bug post-handoff. "
        "Steps 41-46 (master merge) run ONCE per work-item and are not repeated for rework."
    )
    ws["A2"].font = NOTE_FONT
    ws.merge_cells("A2:D2")

    header_row(ws, 4, ["#", "Step", "Phase", "Repeats per rework cycle?"])

    row = 5
    label_map = {
        "pre_qa": ("Dev-test (1-35)", "Yes — every rework cycle"),
        "post_qa": ("QA cycle (36-40)", "Only if QA catches the bug"),
        "master": ("Master merge (41-46)", "No — once per work-item"),
    }
    for idx, desc, status in PROCESS_STEPS:
        ws.cell(row=row, column=1, value=idx).border = BORDER
        ws.cell(row=row, column=2, value=desc).border = BORDER
        phase, repeats = label_map[status]
        ws.cell(row=row, column=3, value=phase).border = BORDER
        ws.cell(row=row, column=4, value=repeats).border = BORDER
        row += 1


# ---------------------------------------------------------------------------
# Defect-Distribution-PCE sheet
# ---------------------------------------------------------------------------

def build_defect_distribution(wb: Workbook):
    ws = wb.create_sheet("Defect-Distribution-PCE")
    widen(ws, {"A": 18, "B": 12, "C": 12, "D": 14, "E": 14, "F": 18, "G": 18})

    ws["A1"] = "Defect Distribution and Phase Containment Effectiveness (PCE)"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:G1")

    header_row(ws, 3, ["Stage", "% Today", "% Target", "Multiplier", "Weighted Today", "Weighted Target", "Cost Delta"])

    stages = [
        ("Local", "pct_today_local", "pct_target_local", "mult_local"),
        ("Development", "pct_today_dev", "pct_target_dev", "mult_dev"),
        ("Staging / UAT", "pct_today_stg", "pct_target_stg", "mult_stg"),
        ("Production", "pct_today_prod", "pct_target_prod", "mult_prod"),
    ]

    row = 4
    first_stage_row = row
    for stage, today, target, mult in stages:
        ws.cell(row=row, column=1, value=stage).border = BORDER
        ws.cell(row=row, column=2, value=f"={today}").number_format = "0.0%"
        ws.cell(row=row, column=3, value=f"={target}").number_format = "0.0%"
        # Multiplier — when stage is Development and toggle ON, swap for bottom-up value.
        if stage == "Development":
            ws.cell(
                row=row,
                column=4,
                value=("=IF(use_bottom_up_multiplier=1,"
                       "development_fix_total/(local_fix_hours*hourly_rate),"
                       "mult_dev)"),
            )
        else:
            ws.cell(row=row, column=4, value=f"={mult}")
        ws.cell(row=row, column=4).number_format = "0.00"
        # Weighted today/target/delta
        ws.cell(row=row, column=5, value=f"=B{row}*D{row}").number_format = "0.00"
        ws.cell(row=row, column=6, value=f"=C{row}*D{row}").number_format = "0.00"
        ws.cell(row=row, column=7, value=f"=E{row}-F{row}").number_format = "0.00"
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = BORDER
        row += 1
    last_stage_row = row - 1

    # Totals row
    total_row(ws, row, cols=7)
    ws.cell(row=row, column=1, value="Weighted average defect cost (units)")
    ws.cell(row=row, column=2, value=f"=SUM(B{first_stage_row}:B{last_stage_row})").number_format = "0.0%"
    ws.cell(row=row, column=3, value=f"=SUM(C{first_stage_row}:C{last_stage_row})").number_format = "0.0%"
    ws.cell(row=row, column=5, value=f"=SUM(E{first_stage_row}:E{last_stage_row})").number_format = "0.00"
    ws.cell(row=row, column=6, value=f"=SUM(F{first_stage_row}:F{last_stage_row})").number_format = "0.00"
    ws.cell(row=row, column=7, value=f"=SUM(G{first_stage_row}:G{last_stage_row})").number_format = "0.00"
    weighted_avg_today_cell = f"E{row}"
    weighted_avg_target_cell = f"F{row}"
    weighted_delta_cell = f"G{row}"
    row += 2

    # PCE summary
    section_row(ws, row, "Phase Containment Effectiveness", cols=7)
    row += 1
    ws.cell(row=row, column=1, value="PCE(local) today")
    ws.cell(row=row, column=2, value=f"=pct_today_local").number_format = "0.0%"
    row += 1
    ws.cell(row=row, column=1, value="PCE(local) target")
    ws.cell(row=row, column=2, value=f"=pct_target_local").number_format = "0.0%"
    row += 1
    ws.cell(row=row, column=1, value="Capers Jones benchmark — median teams")
    ws.cell(row=row, column=2, value=0.55).number_format = "0.0%"
    row += 1
    ws.cell(row=row, column=1, value="Capers Jones benchmark — elite teams")
    ws.cell(row=row, column=2, value=0.85).number_format = "0.0%"
    row += 2

    # Annual cost calc
    section_row(ws, row, "Annual Rework Cost", cols=7)
    row += 1
    ws.cell(row=row, column=1, value="Local fix cost ($)")
    ws.cell(row=row, column=2, value="=local_fix_hours*hourly_rate").number_format = "$#,##0"
    local_cost_cell = f"B{row}"
    row += 1
    ws.cell(row=row, column=1, value="Average defect cost today ($)")
    ws.cell(row=row, column=2, value=f"={local_cost_cell}*{weighted_avg_today_cell}").number_format = "$#,##0"
    avg_today_cell = f"B{row}"
    row += 1
    ws.cell(row=row, column=1, value="Average defect cost target ($)")
    ws.cell(row=row, column=2, value=f"={local_cost_cell}*{weighted_avg_target_cell}").number_format = "$#,##0"
    avg_target_cell = f"B{row}"
    row += 1
    ws.cell(row=row, column=1, value="Annual rework cost today")
    ws.cell(row=row, column=2, value=f"=bugs_per_year*{avg_today_cell}").number_format = "$#,##0"
    annual_today_cell = f"B{row}"
    row += 1
    ws.cell(row=row, column=1, value="Annual rework cost target")
    ws.cell(row=row, column=2, value=f"=bugs_per_year*{avg_target_cell}").number_format = "$#,##0"
    annual_target_cell = f"B{row}"
    row += 1
    total_row(ws, row, cols=7)
    ws.cell(row=row, column=1, value="ANNUAL SHIFT-LEFT SAVINGS")
    ws.cell(row=row, column=2, value=f"={annual_today_cell}-{annual_target_cell}").number_format = "$#,##0"
    # Named range so ROI Summary can pick it up
    dn = DefinedName(name="annual_shift_left_savings", attr_text=f"'Defect-Distribution-PCE'!$B${row}")
    wb.defined_names["annual_shift_left_savings"] = dn


# ---------------------------------------------------------------------------
# Bug-Cost-by-Stage sheet
# ---------------------------------------------------------------------------

def build_bug_cost_by_stage(wb: Workbook):
    ws = wb.create_sheet("Bug-Cost-by-Stage")
    widen(ws, {"A": 24, "B": 14, "C": 14, "D": 50})

    ws["A1"] = "Per-Bug Cost by Detection Stage"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")

    header_row(ws, 3, ["Stage", "Cost ($/bug)", "Multiplier", "Notes"])

    # Anchor — local cost
    ws.cell(row=5, column=1, value="Local")
    ws.cell(row=5, column=2, value="=local_fix_hours*hourly_rate").number_format = "$#,##0"
    ws.cell(row=5, column=3, value=1.0).number_format = "0.00"
    ws.cell(row=5, column=4, value="Baseline; ~55 min of dev time")

    ws.cell(row=6, column=1, value="Development")
    ws.cell(row=6, column=2, value=("=IF(use_bottom_up_multiplier=1,"
                                       "development_fix_total,"
                                       "$B$5*mult_dev)")).number_format = "$#,##0"
    ws.cell(row=6, column=3, value="=B6/$B$5").number_format = "0.00"
    ws.cell(row=6, column=4, value="Bottom-up if toggle ON; IBM 15x if OFF")

    ws.cell(row=7, column=1, value="Staging")
    ws.cell(row=7, column=2, value="=$B$5*mult_stg").number_format = "$#,##0"
    ws.cell(row=7, column=3, value="=mult_stg").number_format = "0.00"
    ws.cell(row=7, column=4, value="IBM SSI 40x")

    ws.cell(row=8, column=1, value="Production")
    ws.cell(row=8, column=2, value="=$B$5*mult_prod").number_format = "$#,##0"
    ws.cell(row=8, column=3, value="=mult_prod").number_format = "0.00"
    ws.cell(row=8, column=4, value="IBM SSI 100x; corroborated by NIST RTI 2002")

    for r in range(5, 9):
        for c in range(1, 5):
            ws.cell(row=r, column=c).border = BORDER


# ---------------------------------------------------------------------------
# Development-Fix-Workflow sheet
# ---------------------------------------------------------------------------

def build_development_fix_workflow(wb: Workbook):
    ws = wb.create_sheet("Development-Fix-Workflow")
    widen(ws, {"A": 38, "B": 14, "C": 12, "D": 14, "E": 14, "F": 30})

    ws["A1"] = "Development-Fix Workflow — Bottom-Up Cost Per Bug"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:F1")

    header_row(ws, 3, ["Step", "Time (min)", "People", "Idle factor", "Cost ($)", "Notes"])

    # Each row: (step name, time-formula, people, idle-factor-formula, notes)
    # cost = time/60 * hourly_rate * people * (1 if idle factor not applicable else factor adjustment)
    # For idle steps, cost = (time/60) * hourly_rate * people * idle_factor
    # For active steps, idle_factor = 1.0
    steps = [
        ("1. Triage / create work item", "=triage_min", 1, "1", "Often + QA on top"),
        ("2. Branch + repro attempt", "=branch_repro_min", 1, "1", "Repro often fails without local env"),
        ("3. Code fix", "=code_fix_min", 1, "1", "Median bug"),
        ("4. AI code-check prompt", "=ai_check_min", 1, "1", "Plus tool token cost (separate line)"),
        ("5. Push wait (CI build/tests)", "=push_wait_min", 1, "=push_idle_factor", "Idle/context-switched"),
        ("6. PR create + description", "=pr_create_min", 1, "1", ""),
        ("7. Reviewer wait (wall-clock)", "=review_wait_hours*60", 1, "=review_wait_idle_factor", "DORA review wait time"),
        ("8. Reviewer review (first pass)", "=reviewer_review_min", 1, "1", "Reviewer side"),
        ("9. Review rounds — dev side", "=review_round_dev_min*review_rounds", 1, "1", ""),
        ("10. Review rounds — reviewer side", "=review_round_reviewer_min*review_rounds", 1, "1", ""),
        ("11. Approve + merge", "=approve_merge_min", 1, "1", ""),
        ("12. Redeploy to integration (wait)", "=redeploy_min", 1, "=redeploy_idle_factor", "Mostly idle"),
        ("13. Re-run regression + reverify", "=reverify_min", 1, "1", ""),
    ]

    row = 4
    first_step_row = row
    for step, time_f, people, idle_f, notes in steps:
        ws.cell(row=row, column=1, value=step)
        ws.cell(row=row, column=2, value=time_f).number_format = "0.0"
        ws.cell(row=row, column=3, value=people)
        ws.cell(row=row, column=4, value=idle_f).number_format = "0.00"
        # cost formula
        ws.cell(row=row, column=5, value=f"=(B{row}/60)*hourly_rate*C{row}*D{row}").number_format = "$#,##0.00"
        ws.cell(row=row, column=6, value=notes).font = NOTE_FONT
        for c in range(1, 7):
            ws.cell(row=row, column=c).border = BORDER
        row += 1
    last_step_row = row - 1

    # AI tool cost line
    ws.cell(row=row, column=1, value="14. AI tool token cost")
    ws.cell(row=row, column=5, value="=ai_check_tool_cost").number_format = "$#,##0.00"
    for c in range(1, 7):
        ws.cell(row=row, column=c).border = BORDER
    ai_tool_row = row
    row += 1

    # Subtotal
    total_row(ws, row, cols=6)
    ws.cell(row=row, column=1, value="Subtotal per fix (no retry)")
    ws.cell(row=row, column=5, value=f"=SUM(E{first_step_row}:E{ai_tool_row})").number_format = "$#,##0.00"
    subtotal_cell = f"E{row}"
    row += 1

    # With retries
    ws.cell(row=row, column=1, value="× retry adjustment (1 + retry_rate)")
    ws.cell(row=row, column=5, value=f"={subtotal_cell}*(1+retry_rate)").number_format = "$#,##0.00"
    with_retry_cell = f"E{row}"
    row += 1

    # Team contention surcharge
    ws.cell(row=row, column=1, value="Team contention surcharge (other devs impeded)")
    ws.cell(row=row, column=5, value="=(team_size-1)*0.25*hourly_rate").number_format = "$#,##0.00"
    contention_cell = f"E{row}"
    row += 1

    # Context-reload cost
    ws.cell(row=row, column=1, value="Original-dev context reload (~22 min)")
    ws.cell(row=row, column=5, value="=(22/60)*hourly_rate").number_format = "$#,##0.00"
    context_cell = f"E{row}"
    row += 1

    # Pipeline compute attribution
    ws.cell(row=row, column=1, value="Pipeline compute (rough avg)")
    ws.cell(row=row, column=5, value="=pipeline_duration_min*pipeline_cost_per_min*(1+retry_rate)").number_format = "$#,##0.00"
    pipeline_cell = f"E{row}"
    row += 1

    # Grand total
    total_row(ws, row, cols=6)
    ws.cell(row=row, column=1, value="GRAND TOTAL per development-stage fix")
    ws.cell(row=row, column=5, value=f"={with_retry_cell}+{contention_cell}+{context_cell}+{pipeline_cell}").number_format = "$#,##0.00"
    grand_total_row = row
    ws.cell(row=row, column=2, value=f"=E{row}").number_format = "$#,##0.00"
    row += 2

    dn = DefinedName(name="development_fix_total", attr_text=f"'Development-Fix-Workflow'!$E${grand_total_row}")
    wb.defined_names["development_fix_total"] = dn

    # Annual workflow tax
    section_row(ws, row, "Annual Workflow Tax", cols=6)
    row += 1
    ws.cell(row=row, column=1, value="Development-stage bugs per year")
    ws.cell(row=row, column=5, value="=bugs_per_year*pct_today_dev").number_format = "#,##0"
    dev_bugs_cell = f"E{row}"
    row += 1
    ws.cell(row=row, column=1, value="Bugs shifted to local (annual)")
    ws.cell(row=row, column=5, value=f"={dev_bugs_cell}*pct_preventable_with_local").number_format = "#,##0"
    shifted_cell = f"E{row}"
    row += 1
    ws.cell(row=row, column=1, value="Per-fix workflow tax (vs local fix cost)")
    ws.cell(row=row, column=5, value=f"=development_fix_total - (local_fix_hours*hourly_rate)").number_format = "$#,##0.00"
    per_fix_tax_cell = f"E{row}"
    row += 1
    total_row(ws, row, cols=6)
    ws.cell(row=row, column=1, value="ANNUAL WORKFLOW TAX RECOVERABLE (if toggle OFF)")
    ws.cell(row=row, column=5, value=f"={shifted_cell}*{per_fix_tax_cell}").number_format = "$#,##0"
    dn2 = DefinedName(name="annual_workflow_tax", attr_text=f"'Development-Fix-Workflow'!$E${row}")
    wb.defined_names["annual_workflow_tax"] = dn2
    row += 2

    # ---- Rework cycles (calibrated from real team data) ----------------------
    section_row(ws, row, "Rework Cycles (calibrated from Real-Bug-Data tab)", cols=6)
    row += 1
    ws.cell(row=row, column=1, value="Annualised rework items (real data)")
    ws.cell(row=row, column=5, value="=real_rework_items_annual").number_format = "#,##0"
    rework_items_cell = f"E{row}"
    row += 1
    ws.cell(row=row, column=1, value="Avg PRs per rework item (real)")
    ws.cell(row=row, column=5, value="=real_avg_prs_per_rework").number_format = "0.00"
    avg_prs_ref = f"E{row}"
    row += 1
    ws.cell(row=row, column=1, value="Extra PR cycles per rework item (avg - 1)")
    ws.cell(row=row, column=5, value=f"={avg_prs_ref}-1").number_format = "0.00"
    extra_per_item_cell = f"E{row}"
    row += 1
    ws.cell(row=row, column=1, value="Blended redo factor")
    ws.cell(row=row, column=5, value="=redo_factor_blended").number_format = "0.000"
    redo_cell = f"E{row}"
    row += 1
    ws.cell(row=row, column=1, value="Cost per extra rework cycle ($)")
    ws.cell(row=row, column=5, value=f"=development_fix_total*{redo_cell}").number_format = "$#,##0.00"
    cost_per_extra_cell = f"E{row}"
    row += 1
    ws.cell(row=row, column=1, value="Annual rework workflow cost (total)")
    ws.cell(
        row=row,
        column=5,
        value=f"={rework_items_cell}*{extra_per_item_cell}*{cost_per_extra_cell}",
    ).number_format = "$#,##0"
    rework_total_cell = f"E{row}"
    row += 1
    total_row(ws, row, cols=6)
    ws.cell(row=row, column=1, value="ANNUAL REWORK-CYCLE TAX RECOVERABLE (real data x preventable)")
    ws.cell(
        row=row,
        column=5,
        value=f"={rework_total_cell}*pct_preventable_with_local",
    ).number_format = "$#,##0"
    dn3 = DefinedName(name="annual_rework_savings", attr_text=f"'Development-Fix-Workflow'!$E${row}")
    wb.defined_names["annual_rework_savings"] = dn3


# ---------------------------------------------------------------------------
# Pipeline-Cost sheet
# ---------------------------------------------------------------------------

def build_pipeline_cost(wb: Workbook):
    ws = wb.create_sheet("Pipeline-Cost")
    widen(ws, {"A": 38, "B": 16, "C": 50})

    ws["A1"] = "Azure DevOps Pipeline Cost Components"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:C1")

    header_row(ws, 3, ["Component", "Annual $", "Notes"])

    row = 4
    section_row(ws, row, "1. Compute Cost (visible Azure invoice)", cols=3)
    row += 1
    ws.cell(row=row, column=1, value="Total annual builds")
    ws.cell(row=row, column=2, value="=team_size*builds_per_dev_per_day*working_days_per_year").number_format = "#,##0"
    annual_builds_cell = f"B{row}"
    row += 1
    ws.cell(row=row, column=1, value="Effective pipeline duration (min)")
    ws.cell(row=row, column=2, value="=effective_pipeline_min").number_format = "0.0"
    ws.cell(row=row, column=3, value="Real-Build-Times median when use_real_data=1; else placeholder").font = NOTE_FONT
    effective_pipe_cell = f"B{row}"
    row += 1
    ws.cell(row=row, column=1, value="Total pipeline minutes/year")
    ws.cell(row=row, column=2, value=f"={annual_builds_cell}*{effective_pipe_cell}").number_format = "#,##0"
    annual_minutes_cell = f"B{row}"
    row += 1
    ws.cell(row=row, column=1, value="Compute cost (per-min)")
    ws.cell(row=row, column=2, value=f"={annual_minutes_cell}*pipeline_cost_per_min").number_format = "$#,##0"
    compute_cost_cell = f"B{row}"
    row += 1
    ws.cell(row=row, column=1, value="Parallel-job subscription")
    ws.cell(
        row=row,
        column=2,
        value=(
            "=IF(use_real_data=1,"
            "(ms_hosted_parallel_jobs*ms_hosted_monthly_cost+self_hosted_parallel_jobs*self_hosted_monthly_cost)*12,"
            "parallel_jobs_extra*parallel_job_monthly_cost*12)"
        ),
    ).number_format = "$#,##0"
    ws.cell(row=row, column=3, value="Split into MS-hosted + self-hosted when real data is on").font = NOTE_FONT
    parallel_jobs_cell = f"B{row}"
    row += 2

    section_row(ws, row, "2. Engineer Idle Time", cols=3)
    row += 1
    ws.cell(row=row, column=1, value="Idle hours/year (queue + build wait)")
    ws.cell(row=row, column=2, value=f"={annual_builds_cell}*(queue_wait_min+{effective_pipe_cell})/60").number_format = "#,##0"
    idle_hours_cell = f"B{row}"
    row += 1
    ws.cell(row=row, column=1, value="Idle cost (with recovery factor)")
    ws.cell(row=row, column=2, value=f"={idle_hours_cell}*hourly_rate*pipeline_idle_recovery_factor").number_format = "$#,##0"
    idle_cost_cell = f"B{row}"
    row += 2

    section_row(ws, row, "3. Failed-Build Retry Overhead", cols=3)
    row += 1
    ws.cell(row=row, column=1, value="Failed builds/year")
    ws.cell(row=row, column=2, value=f"={annual_builds_cell}*pipeline_failure_rate").number_format = "#,##0"
    failed_builds_cell = f"B{row}"
    row += 1
    ws.cell(row=row, column=1, value="Of which preventable by local env")
    ws.cell(row=row, column=2, value=f"={failed_builds_cell}*pct_retries_caused_by_local").number_format = "#,##0"
    preventable_failures_cell = f"B{row}"
    row += 1
    ws.cell(row=row, column=1, value="Cost per retry (compute + idle)")
    ws.cell(
        row=row,
        column=2,
        value=(
            f"=({effective_pipe_cell}*pipeline_cost_per_min)"
            f"+((queue_wait_min+{effective_pipe_cell})/60*hourly_rate*pipeline_idle_recovery_factor)"
        ),
    ).number_format = "$#,##0.00"
    cost_per_retry_cell = f"B{row}"
    row += 1
    ws.cell(row=row, column=1, value="Annual retry cost (recoverable)")
    ws.cell(row=row, column=2, value=f"={preventable_failures_cell}*{cost_per_retry_cell}").number_format = "$#,##0"
    retry_cost_cell = f"B{row}"
    row += 2

    section_row(ws, row, "Summary", cols=3)
    row += 1
    total_row(ws, row, cols=3)
    ws.cell(row=row, column=1, value="Total all-in pipeline cost (today)")
    ws.cell(row=row, column=2, value=f"={compute_cost_cell}+{parallel_jobs_cell}+{idle_cost_cell}+{retry_cost_cell}").number_format = "$#,##0"
    row += 1
    total_row(ws, row, cols=3)
    ws.cell(row=row, column=1, value="ANNUAL PIPELINE-RETRY SAVINGS (recoverable)")
    ws.cell(row=row, column=2, value=f"={retry_cost_cell}").number_format = "$#,##0"
    dn = DefinedName(name="annual_pipeline_savings", attr_text=f"'Pipeline-Cost'!$B${row}")
    wb.defined_names["annual_pipeline_savings"] = dn


# ---------------------------------------------------------------------------
# Cycle-Time-Sprint sheet
# ---------------------------------------------------------------------------

def build_cycle_time_sprint(wb: Workbook):
    ws = wb.create_sheet("Cycle-Time-Sprint")
    widen(ws, {"A": 42, "B": 16, "C": 50})

    ws["A1"] = "Cycle Time and Sprint Capacity Loss"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:C1")

    header_row(ws, 3, ["Metric", "Value", "Notes"])

    row = 4
    section_row(ws, row, "Per-Iteration Cycle Time", cols=3)
    row += 1
    ws.cell(row=row, column=1, value="Cycle time today (min)")
    ws.cell(row=row, column=2, value="=cycle_time_today_min").number_format = "0.0"
    row += 1
    ws.cell(row=row, column=1, value="Cycle time target (min)")
    ws.cell(row=row, column=2, value="=cycle_time_target_min").number_format = "0.0"
    row += 1
    ws.cell(row=row, column=1, value="Delta (min/iteration)")
    ws.cell(row=row, column=2, value="=cycle_time_today_min-cycle_time_target_min").number_format = "0.0"
    row += 2

    section_row(ws, row, "Daily and Sprint Loss Per Dev", cols=3)
    row += 1
    ws.cell(row=row, column=1, value="Daily cycle-time lost (min)")
    ws.cell(row=row, column=2, value="=iterations_per_dev_per_day*(cycle_time_today_min-cycle_time_target_min)").number_format = "#,##0"
    daily_lost_cell = f"B{row}"
    row += 1
    ws.cell(row=row, column=1, value="Daily effective loss (after recovery)")
    ws.cell(row=row, column=2, value=f"={daily_lost_cell}*cycle_idle_recovery_factor/60").number_format = "0.00"
    daily_eff_cell = f"B{row}"
    row += 1
    ws.cell(row=row, column=1, value="Sprint hours lost per dev")
    ws.cell(row=row, column=2, value=f"={daily_eff_cell}*working_days_per_sprint").number_format = "#,##0"
    sprint_per_dev_cell = f"B{row}"
    row += 1
    ws.cell(row=row, column=1, value="Sprint nominal capacity per dev (hrs)")
    ws.cell(row=row, column=2, value="=working_days_per_sprint*7").number_format = "#,##0"
    sprint_nominal_cell = f"B{row}"
    row += 1
    ws.cell(row=row, column=1, value="Sprint capacity erosion %")
    ws.cell(row=row, column=2, value=f"={sprint_per_dev_cell}/{sprint_nominal_cell}").number_format = "0.0%"
    row += 2

    section_row(ws, row, "Annual Team-Wide Impact", cols=3)
    row += 1
    ws.cell(row=row, column=1, value="Annual hours lost to cycle time")
    ws.cell(row=row, column=2, value=f"={sprint_per_dev_cell}*team_size*sprints_per_year").number_format = "#,##0"
    annual_hours_cell = f"B{row}"
    row += 1
    total_row(ws, row, cols=3)
    ws.cell(row=row, column=1, value="ANNUAL SPRINT CAPACITY $ LOST (recoverable)")
    ws.cell(row=row, column=2, value=f"={annual_hours_cell}*hourly_rate*context_switch_multiplier").number_format = "$#,##0"
    dn = DefinedName(name="annual_sprint_savings", attr_text=f"'Cycle-Time-Sprint'!$B${row}")
    wb.defined_names["annual_sprint_savings"] = dn


# ---------------------------------------------------------------------------
# Workarounds sheet
# ---------------------------------------------------------------------------

def build_workarounds(wb: Workbook):
    ws = wb.create_sheet("Workarounds")
    widen(ws, {"A": 4, "B": 38, "C": 14, "D": 14, "E": 18})

    ws["A1"] = "Workarounds — Self-Assessment"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:E1")

    header_row(ws, 3, ["#", "Workaround", "We do this? (Y/P/N)", "Hours/week", "Annual cost"])

    workarounds = [
        "Sharing one remote dev/test environment",
        "Hardcoded conn strings to Azure dev resources",
        "Disabled auth/security locally",
        "Push-to-test debugging (printf via CI)",
        "Mocking Azure services in code instead of emulating",
        "Maintaining personal/shadow Azure subscription",
        "Long-lived feature branches",
        "Comment-out-and-redeploy debugging",
        "Pair-debugging on a shared environment",
        "Skipping integration tests locally",
        "Manual data-prep scripts to seed shared DBs",
        "Long onboarding (1-2 weeks waiting for access)",
        '"Your turn QA" handoffs',
        "Disabled regression suite locally",
    ]

    row = 4
    first_row = row
    for i, w in enumerate(workarounds, start=1):
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=w)
        ws.cell(row=row, column=3, value="").fill = INPUT_FILL  # team fills in
        # Per-row hours: distribute the total across 14 if blank
        ws.cell(row=row, column=4, value=f"=workaround_hours_per_week_total/14").number_format = "0.0"
        ws.cell(row=row, column=5, value=f"=D{row}*workaround_weeks_per_year*hourly_rate").number_format = "$#,##0"
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = BORDER
        row += 1
    last_row = row - 1

    total_row(ws, row, cols=5)
    ws.cell(row=row, column=2, value="ANNUAL WORKAROUND TAX")
    ws.cell(row=row, column=4, value=f"=SUM(D{first_row}:D{last_row})").number_format = "0.0"
    ws.cell(row=row, column=5, value=f"=SUM(E{first_row}:E{last_row})").number_format = "$#,##0"
    dn = DefinedName(name="annual_workaround_savings", attr_text=f"'Workarounds'!$E${row}")
    wb.defined_names["annual_workaround_savings"] = dn


# ---------------------------------------------------------------------------
# ROI-Summary sheet
# ---------------------------------------------------------------------------

def build_roi_summary(wb: Workbook):
    ws = wb.create_sheet("ROI-Summary")
    widen(ws, {"A": 48, "B": 18, "C": 50})

    ws["A1"] = "ROI Summary"
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells("A1:C1")

    header_row(ws, 3, ["Item", "Value", "Notes"])

    row = 4
    section_row(ws, row, "Annual Savings (recovered cost)", cols=3)
    row += 1
    ws.cell(row=row, column=1, value="1. Shift-left rework savings")
    ws.cell(row=row, column=2, value="=annual_shift_left_savings").number_format = "$#,##0"
    p1 = f"B{row}"
    ws.cell(row=row, column=3, value="From Defect-Distribution-PCE tab").font = NOTE_FONT
    row += 1

    ws.cell(row=row, column=1, value="2. Workflow tax (only if toggle OFF; else included in #1)")
    ws.cell(row=row, column=2, value="=IF(use_bottom_up_multiplier=1,0,annual_workflow_tax)").number_format = "$#,##0"
    p2 = f"B{row}"
    ws.cell(row=row, column=3, value="Avoid double-counting").font = NOTE_FONT
    row += 1

    ws.cell(row=row, column=1, value="3. Sprint capacity recovered")
    ws.cell(row=row, column=2, value="=annual_sprint_savings").number_format = "$#,##0"
    p3 = f"B{row}"
    ws.cell(row=row, column=3, value="From Cycle-Time-Sprint tab").font = NOTE_FONT
    row += 1

    ws.cell(row=row, column=1, value="4. Pipeline retries avoided")
    ws.cell(row=row, column=2, value="=annual_pipeline_savings").number_format = "$#,##0"
    p4 = f"B{row}"
    ws.cell(row=row, column=3, value="From Pipeline-Cost tab").font = NOTE_FONT
    row += 1

    ws.cell(row=row, column=1, value="5. Workarounds eliminated")
    ws.cell(row=row, column=2, value="=annual_workaround_savings").number_format = "$#,##0"
    p5 = f"B{row}"
    ws.cell(row=row, column=3, value="From Workarounds tab").font = NOTE_FONT
    row += 1

    ws.cell(row=row, column=1, value="6. Rework cycles avoided (real-data calibrated)")
    ws.cell(row=row, column=2, value="=annual_rework_savings").number_format = "$#,##0"
    p6 = f"B{row}"
    ws.cell(row=row, column=3, value="Extra PR cycles measured in Jan-May 2026 data").font = NOTE_FONT
    row += 1

    total_row(ws, row, cols=3)
    ws.cell(row=row, column=1, value="TOTAL ANNUAL SAVINGS")
    ws.cell(row=row, column=2, value=f"={p1}+{p2}+{p3}+{p4}+{p5}+{p6}").number_format = "$#,##0"
    total_savings_cell = f"B{row}"
    row += 2

    section_row(ws, row, "Investment", cols=3)
    row += 1
    ws.cell(row=row, column=1, value="One-time setup engineer cost")
    ws.cell(row=row, column=2, value="=setup_engineer_hours*hourly_rate").number_format = "$#,##0"
    inv1 = f"B{row}"
    row += 1
    ws.cell(row=row, column=1, value="One-time onboarding/doc cost")
    ws.cell(row=row, column=2, value="=onboarding_doc_hours*hourly_rate").number_format = "$#,##0"
    inv2 = f"B{row}"
    row += 1
    ws.cell(row=row, column=1, value="One-time training cost")
    ws.cell(row=row, column=2, value="=training_hours_per_dev*team_size*hourly_rate").number_format = "$#,##0"
    inv3 = f"B{row}"
    row += 1
    ws.cell(row=row, column=1, value="One-time machine upgrade")
    ws.cell(row=row, column=2, value="=machine_upgrade_cost").number_format = "$#,##0"
    inv4 = f"B{row}"
    row += 1
    total_row(ws, row, cols=3)
    ws.cell(row=row, column=1, value="TOTAL ONE-TIME INVESTMENT")
    ws.cell(row=row, column=2, value=f"={inv1}+{inv2}+{inv3}+{inv4}").number_format = "$#,##0"
    one_time_cell = f"B{row}"
    row += 2

    ws.cell(row=row, column=1, value="Annual Docker license cost")
    ws.cell(row=row, column=2, value="=apply_docker_license_cost*team_size*docker_license_cost_per_user_per_month*12").number_format = "$#,##0"
    ann1 = f"B{row}"
    row += 1
    ws.cell(row=row, column=1, value="Annual maintenance cost")
    ws.cell(row=row, column=2, value="=annual_maintenance_hours*hourly_rate").number_format = "$#,##0"
    ann2 = f"B{row}"
    row += 1
    total_row(ws, row, cols=3)
    ws.cell(row=row, column=1, value="TOTAL ANNUAL INVESTMENT")
    ws.cell(row=row, column=2, value=f"={ann1}+{ann2}").number_format = "$#,##0"
    annual_cost_cell = f"B{row}"
    row += 2

    section_row(ws, row, "ROI", cols=3)
    row += 1
    ws.cell(row=row, column=1, value="Net annual savings (savings − annual cost)")
    ws.cell(row=row, column=2, value=f"={total_savings_cell}-{annual_cost_cell}").number_format = "$#,##0"
    net_cell = f"B{row}"
    row += 1
    ws.cell(row=row, column=1, value="Payback period (months)")
    ws.cell(row=row, column=2, value=f"=IF({net_cell}>0,{one_time_cell}/({net_cell}/12),\"--\")").number_format = "0.0"
    row += 1
    ws.cell(row=row, column=1, value="Year 1 net (savings − one-time − annual)")
    ws.cell(row=row, column=2, value=f"={total_savings_cell}-{one_time_cell}-{annual_cost_cell}").number_format = "$#,##0"
    y1_cell = f"B{row}"
    row += 1
    ws.cell(row=row, column=1, value="Year 2 net")
    ws.cell(row=row, column=2, value=f"={total_savings_cell}-{annual_cost_cell}").number_format = "$#,##0"
    y2_cell = f"B{row}"
    row += 1
    ws.cell(row=row, column=1, value="Year 3 net")
    ws.cell(row=row, column=2, value=f"={total_savings_cell}-{annual_cost_cell}").number_format = "$#,##0"
    y3_cell = f"B{row}"
    row += 1
    ws.cell(row=row, column=1, value="3-year NPV (at discount_rate)")
    ws.cell(row=row, column=2, value=f"={y1_cell}/(1+discount_rate)+{y2_cell}/(1+discount_rate)^2+{y3_cell}/(1+discount_rate)^3").number_format = "$#,##0"
    row += 1
    ws.cell(row=row, column=1, value="3-year ROI (NPV / one-time + 3yr annual)")
    ws.cell(row=row, column=2, value=f"=({y1_cell}+{y2_cell}+{y3_cell})/({one_time_cell}+3*{annual_cost_cell})").number_format = "0.0\"x\""
    row += 2

    section_row(ws, row, "Conservative (Downside) Sensitivity — 25% lower savings", cols=3)
    row += 1
    ws.cell(row=row, column=1, value="Downside annual savings (75%)")
    ws.cell(row=row, column=2, value=f"={total_savings_cell}*0.75").number_format = "$#,##0"
    ds_savings_cell = f"B{row}"
    row += 1
    ws.cell(row=row, column=1, value="Downside net annual")
    ws.cell(row=row, column=2, value=f"={ds_savings_cell}-{annual_cost_cell}").number_format = "$#,##0"
    ds_net_cell = f"B{row}"
    row += 1
    ws.cell(row=row, column=1, value="Downside payback (months)")
    ws.cell(row=row, column=2, value=f"=IF({ds_net_cell}>0,{one_time_cell}/({ds_net_cell}/12),\"--\")").number_format = "0.0"


# ---------------------------------------------------------------------------
# Sensitivity sheet
# ---------------------------------------------------------------------------

def build_sensitivity(wb: Workbook):
    ws = wb.create_sheet("Sensitivity")
    widen(ws, {"A": 36, "B": 18, "C": 18, "D": 18})

    ws["A1"] = "Sensitivity Analysis — How Savings Change with ±25% Inputs"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")

    ws["A3"] = "Modify these scenarios in-place; the ROI Summary will update."
    ws["A3"].font = NOTE_FONT
    ws.merge_cells("A3:D3")

    header_row(ws, 5, ["Input", "Low (-25%)", "Baseline", "High (+25%)"])

    rows = [
        ("Bugs per year", "=bugs_per_year*0.75", "=bugs_per_year", "=bugs_per_year*1.25"),
        ("Hourly rate", "=hourly_rate*0.75", "=hourly_rate", "=hourly_rate*1.25"),
        ("Iterations per dev per day", "=iterations_per_dev_per_day*0.75", "=iterations_per_dev_per_day", "=iterations_per_dev_per_day*1.25"),
        ("Cycle time today (min)", "=cycle_time_today_min*0.75", "=cycle_time_today_min", "=cycle_time_today_min*1.25"),
        ("% preventable with local env", "=pct_preventable_with_local*0.75", "=pct_preventable_with_local", "=MIN(pct_preventable_with_local*1.25,1)"),
        ("Pipeline failure rate", "=pipeline_failure_rate*0.75", "=pipeline_failure_rate", "=pipeline_failure_rate*1.25"),
        ("Pct env/integration related", "=pct_env_or_integration_related*0.75", "=pct_env_or_integration_related", "=pct_env_or_integration_related*1.25"),
    ]
    row = 6
    for label, low, base, high in rows:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=low).number_format = "0.00"
        ws.cell(row=row, column=3, value=base).number_format = "0.00"
        ws.cell(row=row, column=4, value=high).number_format = "0.00"
        for c in range(1, 5):
            ws.cell(row=row, column=c).border = BORDER
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="To run a what-if, change the named-range values on the Inputs tab and observe ROI-Summary recalculate.").font = NOTE_FONT
    ws.merge_cells(f"A{row}:D{row}")


# ---------------------------------------------------------------------------
# CSV fallback
# ---------------------------------------------------------------------------

def write_csv(inp: Inputs, path: str):
    """Write a flat CSV of the inputs and the headline outputs (computed in Python)."""
    # Compute the headline numbers in Python so the CSV has actual values, not formulas.
    local_fix_cost = inp.local_fix_hours * inp.hourly_rate

    weighted_today = (
        inp.pct_today_local * inp.mult_local
        + inp.pct_today_dev * inp.mult_dev
        + inp.pct_today_stg * inp.mult_stg
        + inp.pct_today_prod * inp.mult_prod
    )
    weighted_target = (
        inp.pct_target_local * inp.mult_local
        + inp.pct_target_dev * inp.mult_dev
        + inp.pct_target_stg * inp.mult_stg
        + inp.pct_target_prod * inp.mult_prod
    )
    annual_today = inp.bugs_per_year * local_fix_cost * weighted_today
    annual_target = inp.bugs_per_year * local_fix_cost * weighted_target
    shift_left_savings = annual_today - annual_target

    cycle_delta = inp.cycle_time_today_min - inp.cycle_time_target_min
    daily_lost_min = inp.iterations_per_dev_per_day * cycle_delta
    daily_eff_hours = daily_lost_min * inp.cycle_idle_recovery_factor / 60
    annual_hours_lost = (
        daily_eff_hours * inp.working_days_per_sprint * inp.sprints_per_year * inp.team_size
    )
    sprint_savings = annual_hours_lost * inp.hourly_rate * inp.context_switch_multiplier

    # Effective pipeline duration uses real data when toggle is on.
    pr_counts = [p for _, _, p in REAL_BUG_DATA]
    real_items = len(pr_counts)
    real_avg_prs = sum(pr_counts) / real_items if real_items else 0.0
    build_mins = [m for _, m in REAL_BUILD_TIMES]
    real_median_build = sorted(build_mins)[len(build_mins) // 2] if build_mins else inp.pipeline_duration_min
    real_critical_path = max(build_mins) if build_mins else inp.pipeline_duration_min
    effective_pipeline_min = real_critical_path if inp.use_real_data else inp.pipeline_duration_min

    annual_builds = inp.team_size * inp.builds_per_dev_per_day * inp.working_days_per_year
    failed_builds = annual_builds * inp.pipeline_failure_rate
    preventable_failures = failed_builds * inp.pct_retries_caused_by_local
    cost_per_retry = (
        effective_pipeline_min * inp.pipeline_cost_per_min
        + (inp.queue_wait_min + effective_pipeline_min) / 60
        * inp.hourly_rate * inp.pipeline_idle_recovery_factor
    )
    pipeline_savings = preventable_failures * cost_per_retry

    workaround_savings = (
        inp.workaround_hours_per_week_total * inp.workaround_weeks_per_year * inp.hourly_rate
    )

    # Rework-cycle savings calibrated from real data.
    annual_rework_items = real_items * 12 / inp.data_period_months if inp.data_period_months else 0
    redo_factor_blended = (
        inp.pct_rework_caught_pre_qa * inp.redo_factor_pre_qa
        + (1 - inp.pct_rework_caught_pre_qa) * inp.redo_factor_post_qa
    )
    # development_fix_total (per single PR cycle) — replicate the workflow sum
    workflow_min = (
        inp.triage_min + inp.branch_repro_min + inp.code_fix_min + inp.ai_check_min
        + inp.push_wait_min * inp.push_idle_factor + inp.pr_create_min
        + inp.review_wait_hours * 60 * inp.review_wait_idle_factor
        + inp.reviewer_review_min
        + inp.review_round_dev_min * inp.review_rounds
        + inp.review_round_reviewer_min * inp.review_rounds
        + inp.approve_merge_min
        + inp.redeploy_min * inp.redeploy_idle_factor
        + inp.reverify_min
    )
    workflow_cost = workflow_min / 60 * inp.hourly_rate + inp.ai_check_tool_cost_per_run
    workflow_cost *= (1 + inp.retry_rate)
    workflow_cost += (inp.team_size - 1) * 0.25 * inp.hourly_rate  # contention
    workflow_cost += (22 / 60) * inp.hourly_rate  # context reload
    workflow_cost += effective_pipeline_min * inp.pipeline_cost_per_min * (1 + inp.retry_rate)
    extra_prs_per_item = max(real_avg_prs - 1, 0)
    cost_per_extra_cycle = workflow_cost * redo_factor_blended
    annual_rework_workflow_cost = annual_rework_items * extra_prs_per_item * cost_per_extra_cycle
    rework_savings = annual_rework_workflow_cost * inp.pct_preventable_with_local

    total_annual_savings = (
        shift_left_savings + sprint_savings + pipeline_savings + workaround_savings + rework_savings
    )

    one_time = (
        inp.setup_engineer_hours * inp.hourly_rate
        + inp.onboarding_doc_hours * inp.hourly_rate
        + inp.training_hours_per_dev * inp.team_size * inp.hourly_rate
        + inp.machine_upgrade_cost
    )
    docker_cost = (
        (inp.team_size * inp.docker_license_cost_per_user_per_month * 12)
        if inp.apply_docker_license_cost else 0
    )
    annual_cost = docker_cost + inp.annual_maintenance_hours * inp.hourly_rate
    net_annual = total_annual_savings - annual_cost
    payback_months = (one_time / (net_annual / 12)) if net_annual > 0 else float("inf")

    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Section", "Item", "Value"])
        w.writerow(["Inputs", "Team size", inp.team_size])
        w.writerow(["Inputs", "Hourly rate", f"${inp.hourly_rate:,.2f}"])
        w.writerow(["Inputs", "Bugs per year", inp.bugs_per_year])
        w.writerow(["Inputs", "PCE(local) today", f"{inp.pct_today_local:.0%}"])
        w.writerow(["Inputs", "PCE(local) target", f"{inp.pct_target_local:.0%}"])
        w.writerow(["Inputs", "Cycle time today (min)", inp.cycle_time_today_min])
        w.writerow(["Inputs", "Cycle time target (min)", inp.cycle_time_target_min])
        w.writerow(["Computed", "Weighted defect cost today (units)", f"{weighted_today:.2f}"])
        w.writerow(["Computed", "Weighted defect cost target (units)", f"{weighted_target:.2f}"])
        w.writerow(["Computed", "Local fix cost ($)", f"${local_fix_cost:,.2f}"])
        w.writerow(["Real Data", "Rework items observed (4.25 mo)", real_items])
        w.writerow(["Real Data", "Avg PRs per rework item", f"{real_avg_prs:.2f}"])
        w.writerow(["Real Data", "Annualised rework items", f"{annual_rework_items:,.0f}"])
        w.writerow(["Real Data", "Blended redo factor", f"{redo_factor_blended:.3f}"])
        w.writerow(["Real Data", "Median build duration (min)", f"{real_median_build:.1f}"])
        w.writerow(["Real Data", "Critical-path build duration (min)", f"{real_critical_path:.1f}"])
        w.writerow(["Real Data", "Effective pipeline duration used (min)", f"{effective_pipeline_min:.1f}"])
        w.writerow(["Savings", "1. Shift-left rework", f"${shift_left_savings:,.0f}"])
        w.writerow(["Savings", "2. Sprint capacity", f"${sprint_savings:,.0f}"])
        w.writerow(["Savings", "3. Pipeline retries", f"${pipeline_savings:,.0f}"])
        w.writerow(["Savings", "4. Workarounds", f"${workaround_savings:,.0f}"])
        w.writerow(["Savings", "5. Rework cycles (real data)", f"${rework_savings:,.0f}"])
        w.writerow(["Savings", "TOTAL ANNUAL SAVINGS", f"${total_annual_savings:,.0f}"])
        w.writerow(["Investment", "One-time investment", f"${one_time:,.0f}"])
        w.writerow(["Investment", "Annual ongoing cost", f"${annual_cost:,.0f}"])
        w.writerow(["ROI", "Net annual savings", f"${net_annual:,.0f}"])
        w.writerow(["ROI", "Payback period (months)", f"{payback_months:.1f}" if payback_months != float("inf") else "n/a"])

    return {
        "shift_left_savings": shift_left_savings,
        "sprint_savings": sprint_savings,
        "pipeline_savings": pipeline_savings,
        "workaround_savings": workaround_savings,
        "rework_savings": rework_savings,
        "total_annual_savings": total_annual_savings,
        "one_time": one_time,
        "annual_cost": annual_cost,
        "net_annual": net_annual,
        "payback_months": payback_months,
        "real_items_observed": real_items,
        "real_avg_prs_per_item": real_avg_prs,
        "real_median_build_min": real_median_build,
    }


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    inp = Inputs()
    wb = build_workbook(inp)
    wb.active = wb["ROI-Summary"]
    wb.save("cost-model.xlsx")
    print("Wrote cost-model.xlsx")

    summary = write_csv(inp, "cost-model.csv")
    print("Wrote cost-model.csv")
    print()
    print("--- Headline Outputs (placeholder defaults) ---")
    for k, v in summary.items():
        if "savings" in k or "cost" in k or "annual" in k or "one_time" in k:
            print(f"  {k:30s}  ${v:>12,.0f}")
        else:
            print(f"  {k:30s}  {v}")


if __name__ == "__main__":
    main()
